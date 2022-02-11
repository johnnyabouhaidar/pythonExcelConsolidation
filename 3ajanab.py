import os
import gc
import time
import pandas as pd
from openpyxl import load_workbook

tic  = time.perf_counter()
inputFolder = "inputFiles"
inputFileContains = "LBSR"
workbooks =[]
sheetslist = ["Claims_LandD","Claims_DSec","Claims_DSec","Claims_Others","Lbt_LandD","Lbt_Dsec","Lbt_Others","Lbt_Dsec_Mat"]
ranges2read = [
                ["X16","Y17",'',pd.DataFrame(),15,23],#first is row second is column(zero index start)
                ["AA16","AF17",'',pd.DataFrame(),15,26],
                ["X20","Y56",'',pd.DataFrame(),19,23],
                ["AA20","AF56",'',pd.DataFrame(),19,26],
              ]
              #predefined list for each range to be consolidated along with the sheetname and empty dataframe and the initial starting row and column when writing the final range

###
###

template_file_location = os.path.join("template","LBSR_template.xlsx")
output_folder = "outputConsolidated"

book = load_workbook(template_file_location)
writer = pd.ExcelWriter(os.path.join(output_folder,"outPut_Consolidated.xlsx"),engine='openpyxl')
writer.book=book
writer.sheets = dict((ws.title,ws) for ws in book.worksheets)


print("Workbooks loading")
for file in os.listdir(inputFolder):
    wb = load_workbook(os.path.join(inputFolder,file),data_only=True)
    workbooks.append(wb)

#print(workbooks)
print("Workbooks loaded")
for sheet in sheetslist:
    for wb in workbooks:
        current_ws = wb[sheet]
        for range in ranges2read:
            current_range = current_ws[range[0]:range[1]]
            data_rows=[]
            for row in current_range:
                data_cols=[]
                for cell in row:
                    data_cols.append(cell.value)
                data_rows.append(data_cols)
            df = pd.DataFrame(data_rows)
            range[3] = range[3].add(df, fill_value=0)
            range[3].to_excel(writer,sheet_name=sheet,startrow=range[4],startcol=range[5],header=False,index=False)
            del current_range
            del df
        del current_ws
    for newrange in ranges2read:
        newrange[3] = pd.DataFrame()        



#get data and consolidate
'''
for file in os.listdir(inputFolder):
    
    if (inputFileContains in file):
        wb1 = load_workbook(os.path.join(inputFolder,file))
        for index, currentsheet in enumerate(sheetslist):
            
            #for newrange in ranges2read:
                #newrange[3] = pd.DataFrame()
                    

            ws1 = wb1[currentsheet]
            
            for range in ranges2read:        
                
                rng = ws1[range[0]:range[1]]
                data_rows=[]
                for row in rng:
                    data_cols=[]
                    for cell in row:
                        data_cols.append(cell.value)
                    data_rows.append(data_cols)
                df = pd.DataFrame(data_rows)
                range[3] = range[3].add(df, fill_value=0)
                range[3].to_excel(writer,sheet_name=currentsheet,startrow=range[4],startcol=range[5],header=False,index=False)
        '''        

print("Saving...")

for wbb in workbooks:
    del wbb


del ranges2read
del book

del workbooks

writer.save()
del writer
print("Done Saving!")

toc  = time.perf_counter()



print(f"Done in {toc - tic:0.4f} seconds")


gc.collect()

#print(ranges2read)    

   