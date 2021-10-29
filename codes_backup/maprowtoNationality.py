import os
from openpyxl import load_workbook
from openpyxl.utils.cell import coordinate_from_string , column_index_from_string


#load input file and template file in addition to output files location
input_file_location = r'input2_fornationaltiy\input1consolidated.xlsx'
input_sheet_name = 'Sheet1'
empty_template_location = r'template\final_file.xlsx'

wb1 = load_workbook(input_file_location)
ws1 = wb1[input_sheet_name]
rng = ws1['B2':'F28']

wbtemplate = load_workbook(empty_template_location)
wstemplate = wbtemplate[input_sheet_name]
rngtemplate = wstemplate['B3':'B28']

final_file_location = r'mappedOutput\output.xlsx'


#map rows from input 1 to consolidated template
for inputrow in rng:
    for templaterow in rngtemplate:
        if inputrow[0].value == templaterow[0].value and inputrow[0].value is not None:
            #print(templaterow[0].coordinate)
            iter =0
            for cell in inputrow:                
                if iter!=0:
                    #print(coordinate_from_string(cell.coordinate)[0])
                    wstemplate[coordinate_from_string(cell.coordinate)[0]+str(coordinate_from_string(templaterow[0].coordinate)[1])]=cell.value
                iter=iter+1
            break


#save output workbook            
wbtemplate.save(final_file_location)