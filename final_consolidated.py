import os
import pandas as pd
from openpyxl import load_workbook


inputFolder = "inputFiles"
inputFileContains = "LBSR"
sheetslist = ["Claims_LandD","Claims_DSec"]
ranges2read = [

              ]
              #predefined list for each range to be consolidated along with the sheetname and empty dataframe and the initial starting row and column when writing the final range

for sheet in sheetslist:
    ranges2read.append(["X16","Y17",sheet,pd.DataFrame(),15,23])
    ranges2read.append(["AA16","AF17",sheet,pd.DataFrame(),15,26])
    ranges2read.append(["X20","Y56",sheet,pd.DataFrame(),19,23])
    ranges2read.append(["AA20","AF56",sheet,pd.DataFrame(),19,26])
    ranges2read.append(["X16","Y17",sheet,pd.DataFrame(),15,23])
    ranges2read.append(["AA16","AF17",sheet,pd.DataFrame(),15,26])
    ranges2read.append(["X20","Y56",sheet,pd.DataFrame(),19,23])
    ranges2read.append(["AA20","AF56",sheet,pd.DataFrame(),19,26])
    ranges2read.append(["X16","Y17",sheet,pd.DataFrame(),15,23])
    ranges2read.append(["AA16","AF17",sheet,pd.DataFrame(),15,26])
    ranges2read.append(["X20","Y56",sheet,pd.DataFrame(),19,23])
    ranges2read.append(["AA20","AF56",sheet,pd.DataFrame(),19,26])
    ranges2read.append(["X16","Y17",sheet,pd.DataFrame(),15,23])
    ranges2read.append(["AA16","AF17",sheet,pd.DataFrame(),15,26])
    ranges2read.append(["X20","Y56",sheet,pd.DataFrame(),19,23])
    ranges2read.append(["AA20","AF56",sheet,pd.DataFrame(),19,26])
    ranges2read.append(["X16","Y17",sheet,pd.DataFrame(),15,23])
    ranges2read.append(["AA16","AF17",sheet,pd.DataFrame(),15,26])
    ranges2read.append(["X20","Y56",sheet,pd.DataFrame(),19,23])
    ranges2read.append(["AA20","AF56",sheet,pd.DataFrame(),19,26])
    ranges2read.append(["X16","Y17",sheet,pd.DataFrame(),15,23])
    ranges2read.append(["AA16","AF17",sheet,pd.DataFrame(),15,26])
    ranges2read.append(["X20","Y56",sheet,pd.DataFrame(),19,23])
    ranges2read.append(["AA20","AF56",sheet,pd.DataFrame(),19,26])
    ranges2read.append(["X16","Y17",sheet,pd.DataFrame(),15,23])
    ranges2read.append(["AA16","AF17",sheet,pd.DataFrame(),15,26])
    ranges2read.append(["X20","Y56",sheet,pd.DataFrame(),19,23])
    ranges2read.append(["AA20","AF56",sheet,pd.DataFrame(),19,26])
    ranges2read.append(["X16","Y17",sheet,pd.DataFrame(),15,23])
    ranges2read.append(["AA16","AF17",sheet,pd.DataFrame(),15,26])
    ranges2read.append(["X20","Y56",sheet,pd.DataFrame(),19,23])
    ranges2read.append(["AA20","AF56",sheet,pd.DataFrame(),19,26])
    ranges2read.append(["X16","Y17",sheet,pd.DataFrame(),15,23])
    ranges2read.append(["AA16","AF17",sheet,pd.DataFrame(),15,26])
    ranges2read.append(["X20","Y56",sheet,pd.DataFrame(),19,23])
    ranges2read.append(["AA20","AF56",sheet,pd.DataFrame(),19,26])
    ranges2read.append(["X16","Y17",sheet,pd.DataFrame(),15,23])
    ranges2read.append(["AA16","AF17",sheet,pd.DataFrame(),15,26])
    ranges2read.append(["X20","Y56",sheet,pd.DataFrame(),19,23])
    ranges2read.append(["AA20","AF56",sheet,pd.DataFrame(),19,26])
    ranges2read.append(["X16","Y17",sheet,pd.DataFrame(),15,23])
    ranges2read.append(["AA16","AF17",sheet,pd.DataFrame(),15,26])
    ranges2read.append(["X20","Y56",sheet,pd.DataFrame(),19,23])
    ranges2read.append(["AA20","AF56",sheet,pd.DataFrame(),19,26])
    ranges2read.append(["X16","Y17",sheet,pd.DataFrame(),15,23])
    ranges2read.append(["AA16","AF17",sheet,pd.DataFrame(),15,26])
    ranges2read.append(["X20","Y56",sheet,pd.DataFrame(),19,23])
    ranges2read.append(["AA20","AF56",sheet,pd.DataFrame(),19,26])
    ranges2read.append(["X16","Y17",sheet,pd.DataFrame(),15,23])
    ranges2read.append(["AA16","AF17",sheet,pd.DataFrame(),15,26])
    ranges2read.append(["X20","Y56",sheet,pd.DataFrame(),19,23])
    ranges2read.append(["AA20","AF56",sheet,pd.DataFrame(),19,26])
    ranges2read.append(["X16","Y17",sheet,pd.DataFrame(),15,23])
    ranges2read.append(["AA16","AF17",sheet,pd.DataFrame(),15,26])
    ranges2read.append(["X20","Y56",sheet,pd.DataFrame(),19,23])
    ranges2read.append(["AA20","AF56",sheet,pd.DataFrame(),19,26])
    ranges2read.append(["X16","Y17",sheet,pd.DataFrame(),15,23])
    ranges2read.append(["AA16","AF17",sheet,pd.DataFrame(),15,26])
    ranges2read.append(["X20","Y56",sheet,pd.DataFrame(),19,23])
    ranges2read.append(["AA20","AF56",sheet,pd.DataFrame(),19,26])
    ranges2read.append(["X16","Y17",sheet,pd.DataFrame(),15,23])
    ranges2read.append(["AA16","AF17",sheet,pd.DataFrame(),15,26])
    ranges2read.append(["X20","Y56",sheet,pd.DataFrame(),19,23])
    ranges2read.append(["AA20","AF56",sheet,pd.DataFrame(),19,26])
    ranges2read.append(["X16","Y17",sheet,pd.DataFrame(),15,23])
    ranges2read.append(["AA16","AF17",sheet,pd.DataFrame(),15,26])
    ranges2read.append(["X20","Y56",sheet,pd.DataFrame(),19,23])
    ranges2read.append(["AA20","AF56",sheet,pd.DataFrame(),19,26])
    ranges2read.append(["X16","Y17",sheet,pd.DataFrame(),15,23])
    ranges2read.append(["AA16","AF17",sheet,pd.DataFrame(),15,26])
    ranges2read.append(["X20","Y56",sheet,pd.DataFrame(),19,23])
    ranges2read.append(["AA20","AF56",sheet,pd.DataFrame(),19,26])
    ranges2read.append(["X16","Y17",sheet,pd.DataFrame(),15,23])
    ranges2read.append(["AA16","AF17",sheet,pd.DataFrame(),15,26])
    ranges2read.append(["X20","Y56",sheet,pd.DataFrame(),19,23])
    ranges2read.append(["AA20","AF56",sheet,pd.DataFrame(),19,26])
    ranges2read.append(["X16","Y17",sheet,pd.DataFrame(),15,23])
    ranges2read.append(["AA16","AF17",sheet,pd.DataFrame(),15,26])
    ranges2read.append(["X20","Y56",sheet,pd.DataFrame(),19,23])
    ranges2read.append(["AA20","AF56",sheet,pd.DataFrame(),19,26])
    ranges2read.append(["X16","Y17",sheet,pd.DataFrame(),15,23])
    ranges2read.append(["AA16","AF17",sheet,pd.DataFrame(),15,26])
    ranges2read.append(["X20","Y56",sheet,pd.DataFrame(),19,23])
    ranges2read.append(["AA20","AF56",sheet,pd.DataFrame(),19,26])

print(ranges2read)

template_file_location = os.path.join("template","LBSR_template.xlsx")
output_folder = "outputConsolidated"

book = load_workbook(template_file_location)
writer = pd.ExcelWriter(os.path.join(output_folder,"outPut_Consolidated.xlsx"),engine='openpyxl')
writer.book=book
writer.sheets = dict((ws.title,ws) for ws in book.worksheets)

#get data and consolidate
#for file in os.listdir(inputFolder):
#    
#    if (inputFileContains in file):
#        wb1 = load_workbook(os.path.join(inputFolder,file))
#        for index, currentsheet in enumerate(sheetslist):
#            
#            #for newrange in ranges2read:
#                #newrange[3] = pd.DataFrame()
                    

#            ws1 = wb1[currentsheet]
            
#            for range in ranges2read:        
                
 #               rng = ws1[range[0]:range[1]]
 #               data_rows=[]
  #              for row in rng:
   #                 data_cols=[]
    #                for cell in row:
     #                   data_cols.append(cell.value)
      #              data_rows.append(data_cols)
       #         df = pd.DataFrame(data_rows)
        #        range[3] = range[3].add(df, fill_value=0)
         #       range[3].to_excel(writer,sheet_name=currentsheet,startrow=range[4],startcol=range[5],header=False,index=False)
                


#writer.save()
#print(ranges2read)    

   