from openpyxl import load_workbook

wb = load_workbook("tmptmptmp.xlsx",data_only=True)
ws = wb["Sheet1"]

print(ws["E11"].value)