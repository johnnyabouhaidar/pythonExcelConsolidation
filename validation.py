import pandas as pd
from openpyxl import load_workbook



inputFile = "validation_folder_input\M1_IBS_ItemNames.xlsx"
sheetname="Sheet1"

referenceExcelFile = "valid\\newBM.xlsx"
bm_sheetname="Sheet1"

cells_to_check=["E13","F13","G13","H13",
                "E16","F16","G16","H16",
                "E17","F17","G17","H17",
                "E33","F33","G33","H33",
                "E35","F35","G35","H35",]

final_cells_to_check =  ["E11","F11","G11","H11",
                         "E15","F15","G15","H15",
                         "E33","F33","G33","H33",]      

IBS_mapping_cells = [
                    ['C12','Claims_Total','F17'],
                    ['D12','Claims_Total','D17'],
                    ['E12','Claims_Total','E17'],
                    ['C15','Claims_Total','F16'],
                    ['D15','Claims_Total','D16'],
                    ['E15','Claims_Total','E16'],
                    ['C16','Claims_Total','F18'],
                    ['D16','Claims_Total','D18'],
                    ['E16','Claims_Total','E18'],
                    
                    ['C31','Lbt_Total','F17'],
                    ['D31','Lbt_Total','D17'],
                    ['E31','Lbt_Total','E17'],
                    ['C34','Lbt_Total','F16'],
                    ['D34','Lbt_Total','D16'],
                    ['E34','Lbt_Total','E16'],
                    ['C35','Lbt_Total','F18'],
                    ['D35','Lbt_Total','D18'],
                    ['E35','Lbt_Total','E18'],
]

validationtemplate = "validation_folder_input\\validation_for_IBS.xlsx"
lbsr_path = "validation_folder_input\\LBSR.xlsx"



def load_bm_file():
    df = pd.read_excel(referenceExcelFile)
    return df


def get_cell_value(initial_cell_val,data_frame):
    returned_value = "empty"
    for index, vall in data_frame.iterrows():
        if initial_cell_val == vall['item']:
            returned_value=vall['ValueNumeric']
            break

    return returned_value    

def check_file_if_valid(lbsr,validationtemplate):
    wb_filetoVerify = load_workbook(lbsr,data_only=True)
    
    wb_validation_template = load_workbook(validationtemplate)
    ws_validationIBS = wb_validation_template["IBS"]
    
    for item in IBS_mapping_cells:
        ws_filetoverfiy = wb_filetoVerify[item[1]]
        ws_validationIBS[item[0]]=ws_filetoverfiy[item[2]].value
    
    ws_validationIBS = wb_validation_template["M1 domestic "]
    bmDF = load_bm_file()
    
    for cell in cells_to_check:
        value_toWrite = get_cell_value(ws_validationIBS[cell].value,bmDF)
        if value_toWrite !="empty":
            ws_validationIBS[cell] = value_toWrite



    wb_validation_template.save("tmptmptmp.xlsx")
     

    #wb_modified = load_workbook("tmptmptmp.xlsx",data_only=True)
    #ws_modified = wb_modified[sheetname]
    #for finalcells in final_cells_to_check:
    #    print(ws_modified[finalcells].value)



if __name__ == '__main__':
    check_file_if_valid(lbsr_path,validationtemplate)


