import os
import pandas as pd
from openpyxl import load_workbook


inputFolder = "inputFiles"
inputFileContains = "file"
ranges2read = [
                ["B4","H14",'Sheet1',pd.DataFrame(),3,1],#first is row second is column(zero index start)
                ["K4","L14",'Sheet1',pd.DataFrame(),3,10],
                ["D3","D14",'MAIN',pd.DataFrame(),2,3]
              ]
              #predefined list for each range to be consolidated along with the sheetname and empty dataframe and the initial starting row and column when writing the final range

template_file_location = os.path.join("template","template.xlsx")
output_folder = "outputConsolidated"

#get data and consolidate
for file in os.listdir(inputFolder):
    
    if (inputFileContains in file):
        
        wb1 = load_workbook(os.path.join(inputFolder,file))
        
        for range in ranges2read:        
            ws1 = wb1[range[2]]
            rng = ws1[range[0]:range[1]]
            data_rows=[]
            for row in rng:
                data_cols=[]
                for cell in row:
                    data_cols.append(cell.value)
                data_rows.append(data_cols)
            df = pd.DataFrame(data_rows)
            range[3] = range[3].add(df, fill_value=0)


#save consolidated tables

book = load_workbook(template_file_location)
writer = pd.ExcelWriter(os.path.join(output_folder,"outPut_Consolidated.xlsx"),engine='openpyxl')
writer.book=book
writer.sheets = dict((ws.title,ws) for ws in book.worksheets)
for newrange in ranges2read:
    newrange[3].to_excel(writer,sheet_name=newrange[2],startrow=newrange[4],startcol=newrange[5],header=False,index=False)
    writer.save()    

   